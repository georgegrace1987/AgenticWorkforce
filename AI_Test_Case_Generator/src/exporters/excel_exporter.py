from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment

from src.models.requirement_models import RequirementPackage
from src.models.scenario_models import ScenarioPackage
from src.models.testcase_models import TestCasePackage


def export_test_assets(
    output_dir: Path,
    requirements: RequirementPackage,
    scenarios: ScenarioPackage,
    test_cases: TestCasePackage,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "test_cases_export.xlsx"

    req_rows = [r.model_dump() for r in requirements.requirements]
    scenario_rows = [s.model_dump() for s in scenarios.scenarios]
    testcase_rows = []
    for case in test_cases.test_cases:
        row = case.model_dump()
        row["test_steps"] = "\n".join(
            f"{step.step_number}. {step.action} => {step.expected}" for step in case.test_steps
        )
        row["preconditions"] = " | ".join(case.preconditions or [])
        row["postconditions"] = " | ".join(case.postconditions or [])
        testcase_rows.append(row)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(req_rows).to_excel(writer, index=False, sheet_name="Requirements")
        pd.DataFrame(scenario_rows).to_excel(writer, index=False, sheet_name="Scenarios")
        testcase_df = pd.DataFrame(testcase_rows)
        testcase_df.to_excel(writer, index=False, sheet_name="TestCases")

        _format_test_steps_column(writer, testcase_df)

    return output_path


def _format_test_steps_column(writer: pd.ExcelWriter, testcase_df: pd.DataFrame) -> None:
    """Enable wrap text on the test_steps column so each step renders on its own line."""
    if "test_steps" not in testcase_df.columns:
        return

    worksheet = writer.sheets["TestCases"]
    col_index = testcase_df.columns.get_loc("test_steps") + 1  # 1-based for openpyxl
    column_letter = worksheet.cell(row=1, column=col_index).column_letter

    worksheet.column_dimensions[column_letter].width = 60

    wrap_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=col_index)
        cell.alignment = wrap_alignment
        step_count = str(cell.value or "").count("\n") + 1
        worksheet.row_dimensions[row_idx].height = max(15, step_count * 15)
