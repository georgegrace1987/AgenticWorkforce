"""
Create a sample Excel file with test case template.
This helps users understand the expected format.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Define headers
headers = [
    "test_case_id",
    "title",
    "description",
    "test_steps",
    "expected_result",
    "preconditions",
    "postconditions",
    "automation_candidate",
    "test_type",
    "priority"
]

# Add headers with formatting
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 12

# Add sample data
sample_data = [
    [
        "TC001",
        "User Login with Valid Credentials",
        "Test user login with valid email and password",
        "1. Go to login page; 2. Enter email; 3. Enter password; 4. Click login",
        "User is successfully logged in and redirected to dashboard",
        "Browser is open; Test user account exists",
        "User is logged out after test",
        "Yes",
        "Functional",
        "High"
    ],
    [
        "TC002",
        "User Login with Invalid Password",
        "Test user login fails with wrong password",
        "1. Go to login page; 2. Enter email; 3. Enter wrong password; 4. Click login",
        "Error message 'Invalid credentials' is displayed",
        "Browser is open",
        "Login form is cleared",
        "Yes",
        "Functional",
        "High"
    ],
    [
        "TC003",
        "Check Password Field Masking",
        "Verify that password field masks the input",
        "1. Go to login page; 2. Enter text in password field; 3. Verify text is masked",
        "Password input is masked with asterisks or dots",
        "Browser is open",
        "Page is closed",
        "Yes",
        "UI",
        "Medium"
    ]
]

# Add sample rows
for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# Set row heights
ws.row_dimensions[1].height = 25
for row_num in range(2, len(sample_data) + 2):
    ws.row_dimensions[row_num].height = 50

# Freeze the header row
ws.freeze_panes = "A2"

# Save the file
output_path = Path(__file__).parent / "TestCases" / "sample_test_cases.xlsx"
output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)

print(f"Sample Excel file created: {output_path}")
print("\nColumns:")
for i, header in enumerate(headers, 1):
    print(f"  {i}. {header}")
